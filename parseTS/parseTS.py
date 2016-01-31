from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s-%(levelname)s-%(message)s')
logging.debug('Start of program')

wb = load_workbook('Timesheet - Jim Morrison - WE 2016-01-10.xlsx')

sheet_names = wb.get_sheet_names();

print(sheet_names)
if (sheet_names[0] != 'Timesheet'):
  exit()

ws = wb.get_sheet_by_name('Timesheet')

#print(ws.title)
#print(ws['A3'], ws['A3'].value) # None
#print(ws['B3'], ws['B3'].value) # None
#print(ws['C3'], ws['C3'].value) # Name
#print(ws['D3'], ws['D3'].value) # Actual Name
#print(ws['E3'], ws['E3'].value)
#print(ws['F3'], ws['F3'].value)
#print(ws['G3'], ws['G3'].value)

#for i in range(1,8):
#  for j in range(1,8):
#    print(ws.cell(row=i,column=j), ws.cell(row=i,column=j).value)

wsRow     = 7
sundayFlg = False
sundayCnt = 0;
blankFlg  = False
curDay    = None

while True:
  day = ws.cell(row=wsRow,column=4).value
  if (day == None):
    day = ''
  else:
    curDay = day
  curDay = curDay.ljust(10)

  partA  = ws.cell(row=wsRow,column=5).value
  if (partA == None):
    partA = ''
    blankFlg = True
  partA = partA.ljust(20)

  partB = ws.cell(row=wsRow,column=6).value
  if (partB == None):
    partB = ''
  partB = partB.ljust(30)

  partC = ws.cell(row=wsRow,column=7).value
  if (partC == None):
    partC = ''
  partC = partC.ljust(40)

  partD = ws.cell(row=wsRow,column=8).value
  if (partD == None):
    partD = ''
  partD = partD.ljust(20)

  if (not blankFlg):
    logging.debug(curDay + '|' + partA + '|' + partB + '|' + partC + '|' + '|')

  if (day.startswith('Sunday')):
    sundayFlg = True
  if (sundayFlg):
    if (blankFlg):
      sundayCnt += 1
    else:
      sundayCnt = 0;
  if (sundayCnt > 5):
    break
  wsRow += 1
  blankFlg = False

log.debug('Done')
#while True:
#  print('x to exit')
#  inp = input()
#  if inp == 'x':
#    exit()


